import io
import json
import re
import time
import zipfile
import xml.etree.ElementTree as ET
from textwrap import dedent
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from difflib import SequenceMatcher

# ─────────────────────────────────────────────────────────────────────
# Page meta + theming
# ─────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Masterfile Automation - Amazon", page_icon="🧾", layout="wide")
st.markdown("""
<style>
:root{ --bg1:#f6f9fc; --bg2:#fff; --card:#fff; --card-border:#e8eef6;
--ink:#0f172a; --muted:#64748b; --accent:#2563eb; }
.stApp{background:linear-gradient(180deg, var(--bg1) 0%, var(--bg2) 70%);}
.block-container{padding-top:.75rem;}
.section{border:1px solid var(--card-border);background:var(--card);border-radius:16px;
  padding:18px 20px; box-shadow:0 6px 24px rgba(2,6,23,.05); margin-bottom:18px;}
.badge{display:inline-block;padding:4px 10px;border-radius:999px;font-size:.82rem;font-weight:600;margin-right:.25rem}
.badge-info{background:#eef2ff;color:#1e40af} .badge-ok{background:#ecfdf5;color:#065f46}
div.stButton>button,.stDownloadButton>button{background:var(--accent)!important;color:#fff!important;border-radius:10px!important;border:0!important}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────
# Template layout constants
# ─────────────────────────────────────────────────────────────────────
MASTER_TEMPLATE_SHEET = "Template"   # target sheet
MASTER_DISPLAY_ROW    = 2            # human headers
MASTER_SECONDARY_ROW  = 3            # bullet disambiguators
MASTER_DATA_START_ROW = 4            # first data row

# ─────────────────────────────────────────────────────────────────────
# Pre-defined attribute mappings
# ─────────────────────────────────────────────────────────────────────
PREDEFINED_MAPPING = {
    "Product Type": ["Amazon Product Type"],
    "Seller SKU": ["Pattern SKU", "Seller SKU", "item_sku", "SKU"],
    "Product ID": ["UPC", "UPC/EAN", "Product ID", "external_product_id", "barcode", "barcode.value"],
    "Brand Name": ["Brand Name", "Brand", "brand_name", "Walmart Brand Name - en-US"],
    "Product ID Type": ["Product ID Type"],
    "Item Type Keyword": ["Amazon Category (Tree)"],
    "Your Price": ["List Price", "Selling Price"],
    "Manufacturer": ["Manufacturer Name", "Manufacturer Name (for enforcement)", "Manufacturer"],
    "Manufacturer Part Number": ["Manufacturer Part Number"],
    "Description": ["Product Description", "Description", "long_description", "Walmart Description - en-US"],
    "Main Image URL": ["Main Image", "Main Image URL", "main image url", "image url", "Walmart Main Image URL - en-US", "main_image_url - en-US"],
    "Package Length": ["Package Length (IN)", "Packaged Length (IN)", "Package Length", "package_length - en-US", "length", "depth"],
    "Package Height": ["Package Height (IN)", "Packaged Height (IN)", "Package Height", "package_height - en-US", "height"],
    "Package Weight": ["Package Weight (LB)", "Packaged Weight (LB)", "Package Weight", "package_weight - en-US", "weight"],
    "Package Width": ["Package Width (IN)", "Packaged Width (IN)", "Package Width", "package_width - en-US", "width"],
    "Product Name": ["Product Name", "item_name", "Item Name", "Walmart Title - en-US", "Title"],
    "Item Length": ["Item Length (IN)", "Item Length", "item_length - en-US", "length", "depth"],
    "Item Height": ["Item Height (IN)", "Item Height", "item_height - en-US", "height"],
    "Item Weight": ["Item Weight (LB)", "Item Weight", "item_weight - en-US", "weight"],
    "Item Width": ["Item Width (IN)", "Item Width", "item_width - en-US", "width"],
    "Unit Count": ["Measurement Value"],
    "Ingredients": ["Ingredients"],
    "Country of Origin": ["Country of Origin"],
    "Variation Theme": ["Variation Theme"],
    "Directions": ["Directions"],
    "Indications": ["Indications"],
    "Is the Item Heat Sensitive?": ["Is the item heat sensitive? (Y/N)"],
    "List Price": ["List Price", "Selling Price"],
    "Item Form": ["Item Form (Capsule, Softgel, Powder, Tablet, etc)"],
    "Is Product Expirable": ["Able to Expire? (Y/N)"],
    "Fulfillment Center Shelf Life": ["Shelf Life"],
    "Product Expiration Type": ["Shelf Life"],
    "currency": ["List Price Currency Type (USD, EUR, CAD, etc)"],
    "safety_data_sheet_url": ["SDS Sheet"],
    "Other Image URL1": ["Additional Image 1", "Other Image URL1", "Other Image URL 1", "Image 1", "Walmart Additional Image URL #2 - en-US"],
    "Other Image URL2": ["Additional Image 2", "Other Image URL2", "Other Image URL 2", "Image 2", "Walmart Additional Image URL #3 - en-US"],
    "Other Image URL3": ["Additional Image 3", "Other Image URL3", "Other Image URL 3", "Image 3", "Walmart Additional Image URL #4 - en-US"],
    "Other Image URL4": ["Additional Image 4", "Other Image URL4", "Other Image URL 4", "Image 4", "Walmart Additional Image URL #5 - en-US"],
    "Other Image URL5": ["Additional Image 5", "Other Image URL5", "Other Image URL 5", "Image 5", "Walmart Additional Image URL #6 - en-US"],
    "Other Image URL6": ["Additional Image 6", "Other Image URL6", "Other Image URL 6", "Image 6", "Walmart Additional Image URL #7 - en-US"],
    "Other Image URL7": ["Additional Image 7", "Other Image URL7", "Other Image URL 7", "Image 7", "Walmart Additional Image URL #8 - en-US"],
    "Other Image URL8": ["Additional Image 8", "Other Image URL8", "Other Image URL 8", "Image 8", "Walmart Additional Image URL #9 - en-US"],
    "Other Image URL9": ["Additional Image 9", "Other Image URL9", "Other Image URL 9", "Image 9", "Walmart Additional Image URL #10 - en-US"],
    "bullet_point1": ["Bullet point 1", "bullet_point1", "Bullet Feature 1", "bullet point 1", "bullet_point1 - en-US", "Key Features #1 - en-US"],
    "bullet_point2": ["Bullet point 2", "bullet_point2", "Bullet Feature 2", "bullet point 2", "bullet_point2 - en-US", "Key Features #2 - en-US"],
    "bullet_point3": ["Bullet point 3", "bullet_point3", "Bullet Feature 3", "bullet point 3", "bullet_point3 - en-US", "Key Features #3 - en-US"],
    "bullet_point4": ["Bullet point 4", "bullet_point4", "Bullet Feature 4", "bullet point 4", "bullet_point4 - en-US", "Key Features #4 - en-US"],
    "bullet_point5": ["Bullet point 5", "bullet_point5", "Bullet Feature 5", "bullet point 5", "bullet_point5 - en-US", "Key Features #5 - en-US"]
}

# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────
XL_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XL_NS_REL  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", XL_NS_MAIN)
ET.register_namespace("r", XL_NS_REL)
ET.register_namespace("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006")
ET.register_namespace("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")

_INVALID_XML_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\uD800-\uDFFF]")

def sanitize_xml_text(s) -> str:
    if s is None: return ""
    return _INVALID_XML_CHARS.sub("", str(s))

def norm(s: str) -> str:
    if s is None: return ""
    x = str(s).strip().lower()
    x = re.sub(r"\s*-\s*en\s*[-_ ]\s*us\s*$", "", x)
    x = x.replace("–","-").replace("—","-").replace("−","-")
    x = re.sub(r"[._/\\-]+", " ", x)
    x = re.sub(r"[^0-9a-z\s]+", " ", x)
    return re.sub(r"\s+", " ", x).strip()

def top_matches(query, candidates, k=3):
    q = norm(query)
    scored = [(SequenceMatcher(None, q, norm(c)).ratio(), c) for c in candidates]
    scored.sort(key=lambda t: t[0], reverse=True)
    return scored[:k]

def nonempty_rows(df: pd.DataFrame) -> int:
    if df.empty: return 0
    return df.replace("", pd.NA).dropna(how="all").shape[0]

def worksheet_used_cols(ws, header_rows=(1,), hard_cap=2048, empty_streak_stop=8):
    max_col = ws.max_column if ws.max_column is not None else 1
    max_try = min(max_col, hard_cap)
    last_nonempty, streak = 0, 0
    for c in range(1, max_try + 1):
        any_val = any((ws.cell(row=r, column=c).value not in (None, "")) for r in header_rows)
        if any_val: last_nonempty, streak = c, 0
        else:
            streak += 1
            if streak >= empty_streak_stop: break
    return max(last_nonempty, 1)

def _col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n-1, 26)
        s = chr(65+r) + s
    return s

def _col_number(letters: str) -> int:
    n = 0
    for ch in letters:
        if not ch.isalpha(): break
        n = n * 26 + (ord(ch.upper()) - 64)
    return n

# NEW: safe output filename (keeps letters, numbers, space, _ . -)
def safe_filename(name: str, fallback: str = "final_masterfile") -> str:
    if name is None:
        return fallback
    name = name.strip()
    name = re.sub(r"[^A-Za-z0-9._ -]+", "", name)
    return name or fallback

# ── ZIP / XML helpers ────────────────────────────────────────────────
def _find_sheet_part_path(z: zipfile.ZipFile, sheet_name: str) -> str:
    wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid = None
    for sh in wb_xml.find(f"{{{XL_NS_MAIN}}}sheets"):
        if sh.attrib.get("name") == sheet_name:
            rid = sh.attrib.get(f"{{{XL_NS_REL}}}id")
            break
    if not rid: raise ValueError(f"Sheet '{sheet_name}' not found.")
    target = None
    for rel in rels_xml:
        if rel.attrib.get("Id") == rid:
            target = rel.attrib.get("Target")
            break
    if not target: raise ValueError(f"Relationship for sheet '{sheet_name}' not found.")
    target = target.replace("\\", "/")
    if target.startswith("../"): target = target[3:]
    if not target.startswith("xl/"): target = "xl/" + target
    return target  # e.g., xl/worksheets/sheet1.xml

def _get_table_paths_for_sheet(z: zipfile.ZipFile, sheet_path: str) -> list:
    rels_path = sheet_path.replace("worksheets/", "worksheets/_rels/").replace(".xml", ".xml.rels")
    if rels_path not in z.namelist(): return []
    root = ET.fromstring(z.read(rels_path))
    out = []
    for rel in root:
        t = rel.attrib.get("Type", "")
        if t.endswith("/table"):
            target = rel.attrib.get("Target", "").replace("\\", "/")
            if target.startswith("../"): target = target[3:]
            if not target.startswith("xl/"): target = "xl/" + target
            out.append(target)
    return out

def _read_table_cols_count(table_xml_bytes: bytes) -> int:
    try:
        root = ET.fromstring(table_xml_bytes)
        tcols = root.find(f"{{{XL_NS_MAIN}}}tableColumns")
        if tcols is None: return 0
        cnt_attr = tcols.attrib.get("count")
        cnt = int(cnt_attr) if cnt_attr else 0
        child_count = sum(1 for _ in tcols)
        return max(cnt, child_count)
    except Exception:
        return 0

def _union_dimension(orig_dim_ref: str, used_cols: int, last_row: int) -> str:
    try:
        _, right = orig_dim_ref.split(":", 1)
        m = re.match(r"([A-Z]+)(\d+)", right)
        if m:
            orig_last_col = _col_number(m.group(1))
            orig_last_row = int(m.group(2))
        else:
            orig_last_col, orig_last_row = used_cols, last_row
    except Exception:
        orig_last_col, orig_last_row = used_cols, last_row
    u_last_col = max(orig_last_col, used_cols)
    u_last_row = max(orig_last_row, last_row)
    return f"A1:{_col_letter(u_last_col)}{u_last_row}"

def _ensure_ws_x14ac(root):
    # Allow x14ac attributes without repairs
    root.set("{http://schemas.openxmlformats.org/markup-compatibility/2006}Ignorable", "x14ac")

def _intersects_range(a1: str, r1: int, r2: int) -> bool:
    # a1 like "A3:B7" → True if overlap with [r1, r2]
    m = re.match(r"^[A-Z]+(\d+):[A-Z]+(\d+)$", a1 or "", re.I)
    if not m:
        return False
    lo = int(m.group(1)); hi = int(m.group(2))
    if lo > hi: lo, hi = hi, lo
    return not (hi < r1 or lo > r2)

def _patch_sheet_xml(sheet_xml_bytes: bytes, header_row: int, start_row: int, used_cols_final: int, block_2d: list) -> bytes:
    root = ET.fromstring(sheet_xml_bytes)
    _ensure_ws_x14ac(root)

    sheetData = root.find(f"{{{XL_NS_MAIN}}}sheetData")
    if sheetData is None:
        sheetData = ET.SubElement(root, f"{{{XL_NS_MAIN}}}sheetData")

    # 1) Remove existing data rows at/after start_row
    for row in list(sheetData):
        try:
            r = int(row.attrib.get("r") or "0")
        except Exception:
            r = 0
        if r >= start_row:
            sheetData.remove(row)

    # 2) Remove mergeCells that intersect our data region to prevent "Repaired Records"
    mergeCells = root.find(f"{{{XL_NS_MAIN}}}mergeCells")
    if mergeCells is not None:
        for mc in list(mergeCells):
            ref = mc.attrib.get("ref", "")
            if _intersects_range(ref, start_row, 1048576):
                mergeCells.remove(mc)
        if len(list(mergeCells)) == 0:
            root.remove(mergeCells)

    # 3) Append dense rows (A..lastCol) using inlineStr (keeps rows visible, no sparse-row repair)
    row_span = f"1:{used_cols_final}" if used_cols_final > 0 else "1:1"
    n_rows = len(block_2d)
    for i in range(n_rows):
        r = start_row + i
        src_row = block_2d[i]
        row_el = ET.Element(f"{{{XL_NS_MAIN}}}row", r=str(r))
        row_el.set("spans", row_span)
        row_el.set("{http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac}dyDescent", "0.25")

        for j in range(used_cols_final):
            val = src_row[j] if j < len(src_row) else ""
            txt = sanitize_xml_text(val) if val else ""
            col = _col_letter(j + 1)
            c = ET.Element(f"{{{XL_NS_MAIN}}}c", r=f"{col}{r}", t="inlineStr")
            is_el = ET.SubElement(c, f"{{{XL_NS_MAIN}}}is")
            t_el = ET.SubElement(is_el, f"{{{XL_NS_MAIN}}}t")
            t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            t_el.text = txt  # empty allowed → row still visible
            row_el.append(c)

        sheetData.append(row_el)

    # 4) Dimension: conservative union with original
    dim = root.find(f"{{{XL_NS_MAIN}}}dimension")
    if dim is None:
        dim = ET.SubElement(root, f"{{{XL_NS_MAIN}}}dimension", ref="A1:A1")
    last_row = max(header_row, start_row + max(0, n_rows) - 1)
    new_ref = _union_dimension(dim.attrib.get("ref", "A1:A1"), used_cols_final, last_row)
    dim.set("ref", new_ref)

    # 5) AutoFilter: only update if one existed originally
    af = root.find(f"{{{XL_NS_MAIN}}}autoFilter")
    if af is not None:
        af.set("ref", f"A{header_row}:{_col_letter(used_cols_final)}{last_row}")

    # 6) Clear filterMode flag if present (prevents repair on changed rows)
    sheetPr = root.find(f"{{{XL_NS_MAIN}}}sheetPr")
    if sheetPr is not None and sheetPr.attrib.get("filterMode"):
        sheetPr.attrib.pop("filterMode", None)

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def _patch_table_xml(table_xml_bytes: bytes, header_row: int, last_row: int, last_col_n: int) -> bytes:
    root = ET.fromstring(table_xml_bytes)
    new_ref = f"A{header_row}:{_col_letter(last_col_n)}{last_row}"
    root.set("ref", new_ref)

    af = root.find(f"{{{XL_NS_MAIN}}}autoFilter")
    if af is None:
        af = ET.SubElement(root, f"{{{XL_NS_MAIN}}}autoFilter")
    af.set("ref", new_ref)

    # Keep tableColumns list as-is; just ensure the 'count' equals the number of children (Excel requirement)
    tcols = root.find(f"{{{XL_NS_MAIN}}}tableColumns")
    if tcols is not None:
        child_count = sum(1 for _ in tcols)
        tcols.set("count", str(child_count))
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def _strip_calcchain_override(ct_bytes: bytes) -> bytes:
    try:
        ns = "http://schemas.openxmlformats.org/package/2006/content-types"
        root = ET.fromstring(ct_bytes)
        ET.register_namespace("", ns)
        for el in list(root):
            if el.tag == f"{{{ns}}}Override" and el.attrib.get("PartName","").lower() == "/xl/calcchain.xml":
                root.remove(el)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)
    except Exception:
        return ct_bytes

def fast_patch_template(master_bytes: bytes, sheet_name: str, header_row: int, start_row: int, used_cols: int, block_2d: list) -> bytes:
    """Ultra-fast writer: swaps only the target sheet XML + syncs tables & filters; removes calcChain."""
    zin = zipfile.ZipFile(io.BytesIO(master_bytes), "r")
    sheet_path = _find_sheet_part_path(zin, sheet_name)
    table_paths = _get_table_paths_for_sheet(zin, sheet_path)

    # Use at least the widest table width (some tables define more columns than headers)
    max_cols = used_cols
    for tp in table_paths:
        try:
            cnt = _read_table_cols_count(zin.read(tp))
            if cnt > max_cols: max_cols = cnt
        except Exception:
            pass

    new_sheet_xml = _patch_sheet_xml(zin.read(sheet_path), header_row, start_row, max_cols, block_2d)

    last_row = max(header_row, start_row + max(0, len(block_2d)) - 1)
    patched_tables = {}
    for tp in table_paths:
        try:
            patched_tables[tp] = _patch_table_xml(zin.read(tp), header_row, last_row, max_cols)
        except Exception:
            pass

    out_bio = io.BytesIO()
    with zipfile.ZipFile(out_bio, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            fn = item.filename
            if fn == sheet_path:
                zout.writestr(item, new_sheet_xml)
            elif fn in patched_tables:
                zout.writestr(item, patched_tables[fn])
            elif fn.lower() == "[content_types].xml":
                zout.writestr(item, _strip_calcchain_override(zin.read(fn)))
            elif fn.lower() == "xl/calcchain.xml":
                # Drop calcChain so Excel rebuilds without 'repair'
                continue
            else:
                zout.writestr(item, zin.read(fn))
    zin.close()
    out_bio.seek(0)
    return out_bio.getvalue()

# ─────────────────────────────────────────────────────────────────────
# Session state initialization
# ─────────────────────────────────────────────────────────────────────
if 'unmapped_attributes' not in st.session_state:
    st.session_state.unmapped_attributes = []
if 'user_mappings' not in st.session_state:
    st.session_state.user_mappings = {}
if 'master_data' not in st.session_state:
    st.session_state.master_data = {}
if 'onboarding_data' not in st.session_state:
    st.session_state.onboarding_data = {}
if 'last_files' not in st.session_state:
    st.session_state.last_files = (None, None)

# ─────────────────────────────────────────────────────────────────────
# UI — inputs
# ─────────────────────────────────────────────────────────────────────
st.title("🧾 Masterfile Automation – Amazon")
st.caption("Ultra-fast writer with smart pre-configured mappings. Just upload files and map any remaining attributes!")

st.markdown("<div class='section'><span class='badge badge-info'>Pre-configured Mappings</span> "
            "<span class='badge badge-ok'>Fast XML Writer</span></div>", unsafe_allow_html=True)

st.markdown("<div class='section'>", unsafe_allow_html=True)
c1, c2 = st.columns([1, 1])
with c1:
    masterfile_file = st.file_uploader("📄 Masterfile Template (.xlsx / .xlsm)", type=["xlsx", "xlsm"], key="master_upload")
with c2:
    onboarding_file = st.file_uploader("🧾 Onboarding (.xlsx)", type=["xlsx"], key="onboard_upload")

st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────
# Main — Automatic Analysis & Interactive Mapping
# ─────────────────────────────────────────────────────────────────────
SENTINEL_LIST = object()

if masterfile_file and onboarding_file:
    # Check if files changed - reset if so
    current_files = (masterfile_file.name, onboarding_file.name)
    if current_files != st.session_state.last_files:
        st.session_state.last_files = current_files
        st.session_state.master_data = {}
        st.session_state.onboarding_data = {}
        st.session_state.unmapped_attributes = []
        st.session_state.user_mappings = {}
    
    # Use pre-defined mapping
    mapping_aliases = {}
    for k, v in PREDEFINED_MAPPING.items():
        aliases = v[:] if isinstance(v, list) else [v]
        if k not in aliases: aliases.append(k)
        mapping_aliases[norm(k)] = aliases

    # Read files if not already done
    if not st.session_state.master_data:
        with st.spinner("⏳ Analyzing files..."):
            # Read template
            masterfile_file.seek(0)
            master_bytes = masterfile_file.read()
            wb_ro = load_workbook(io.BytesIO(master_bytes), read_only=True, data_only=True, keep_links=True)
            if MASTER_TEMPLATE_SHEET not in wb_ro.sheetnames:
                st.error(f"Sheet **'{MASTER_TEMPLATE_SHEET}'** not found in the masterfile.")
                st.stop()
            ws_ro = wb_ro[MASTER_TEMPLATE_SHEET]
            used_cols = worksheet_used_cols(ws_ro, header_rows=(MASTER_DISPLAY_ROW, MASTER_SECONDARY_ROW), hard_cap=2048, empty_streak_stop=8)
            display_headers = [ws_ro.cell(row=MASTER_DISPLAY_ROW, column=c).value or "" for c in range(1, used_cols+1)]
            secondary_headers = [ws_ro.cell(row=MASTER_SECONDARY_ROW, column=c).value or "" for c in range(1, used_cols+1)]
            wb_ro.close()

            # Pick best onboarding sheet
            best_xl = pd.ExcelFile(onboarding_file)
            best, best_score, best_info = None, -1, ""
            for sheet in best_xl.sheet_names:
                try:
                    df = best_xl.parse(sheet_name=sheet, header=0, dtype=str).fillna("")
                    df.columns = [str(c).strip() for c in df.columns]
                except Exception:
                    continue
                header_set = {norm(c) for c in df.columns}
                matches = sum(any(norm(a) in header_set for a in aliases)
                              for aliases in mapping_aliases.values())
                rows = nonempty_rows(df)
                score = matches + (0.01 if rows > 0 else 0.0)
                if score > best_score:
                    best, best_score = (df, sheet), score
                    best_info = f"matched headers: {matches}, non-empty rows: {rows}"
            if best is None:
                st.error("No readable onboarding sheet found.")
                st.stop()
            best_df, best_sheet = best[0], best[1]
            on_df = best_df.fillna("")
            on_df.columns = [str(c).strip() for c in on_df.columns]
            on_headers = list(on_df.columns)

            # Analyze mappings
            series_by_alias = {norm(h): on_df[h] for h in on_headers}
            BULLET_DISP_N = norm("Key Product Features")
            unmapped = []
            mapped_count = 0
            
            # Get list of predefined attribute names (normalized)
            predefined_norms = {norm(k) for k in PREDEFINED_MAPPING.keys()}

            for c, (disp, sec) in enumerate(zip(display_headers, secondary_headers), start=1):
                disp_norm = norm(disp)
                sec_norm = norm(sec)
                if disp_norm == BULLET_DISP_N and sec_norm:
                    effective_header = sec
                    label_for_log = f"{disp} ({sec})"
                else:
                    effective_header = disp
                    label_for_log = disp
                eff_norm = norm(effective_header)
                if not eff_norm:
                    continue
                
                # Check if this attribute is in predefined mappings
                is_predefined = eff_norm in predefined_norms
                
                aliases = mapping_aliases.get(eff_norm, [effective_header])
                resolved = None
                for a in aliases:
                    s = series_by_alias.get(norm(a))
                    if s is not None:
                        resolved = (s, a)
                        break
                
                if resolved is None and disp_norm != norm("Listing Action (List or Unlist)"):
                    # Only add to unmapped list if it's a predefined attribute
                    if is_predefined:
                        suggestions = top_matches(effective_header, on_headers, 3)
                        unmapped.append({
                            'col_num': c,
                            'master_name': label_for_log,
                            'display': disp,
                            'secondary': sec,
                            'suggestions': suggestions
                        })
                
                if resolved is not None or disp_norm == norm("Listing Action (List or Unlist)"):
                    mapped_count += 1

            # Store in session state
            st.session_state.master_data = {
                'bytes': master_bytes,
                'used_cols': used_cols,
                'display_headers': display_headers,
                'secondary_headers': secondary_headers,
                'ext': (Path(masterfile_file.name).suffix or ".xlsx").lower()
            }
            st.session_state.onboarding_data = {
                'df': on_df,
                'headers': on_headers,
                'series_by_alias': series_by_alias,
                'best_sheet': best_sheet
            }
            st.session_state.unmapped_attributes = unmapped
            st.session_state.mapped_count = mapped_count

    # Show mapping status
    st.markdown("<div class='section'>", unsafe_allow_html=True)
    st.markdown("### 📊 Mapping Analysis")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("✅ Auto-Mapped", st.session_state.mapped_count)
    with col2:
        st.metric("❓ Unmapped", len(st.session_state.unmapped_attributes))
    with col3:
        total = st.session_state.mapped_count + len(st.session_state.unmapped_attributes)
        st.metric("📋 Total Attributes", total)
    st.markdown("</div>", unsafe_allow_html=True)

    # Interactive mapping UI for unmapped attributes
    if st.session_state.unmapped_attributes:
        st.markdown("<div class='section'>", unsafe_allow_html=True)
        st.markdown("### 🔗 Map Remaining Attributes")
        st.caption("These predefined attributes couldn't be auto-mapped. Select the corresponding onboarding column for each.")
        
        on_headers = st.session_state.onboarding_data['headers']
        options = ["(Leave Empty)"] + on_headers
        
        for attr in st.session_state.unmapped_attributes:
            col_num = attr['col_num']
            master_name = attr['master_name']
            suggestions = attr['suggestions']
            
            # Auto-select best match (highest score suggestion)
            if suggestions and len(suggestions) > 0:
                best_match = suggestions[0][1]  # Get the column name with highest score
                default_idx = options.index(best_match) if best_match in options else 0
            else:
                default_idx = 0
            
            selected = st.selectbox(
                f"**{master_name}**",
                options,
                index=default_idx,
                key=f"map_{col_num}"
            )
            
            if selected != "(Leave Empty)":
                st.session_state.user_mappings[col_num] = selected
            elif col_num in st.session_state.user_mappings:
                del st.session_state.user_mappings[col_num]
        
        st.markdown("</div>", unsafe_allow_html=True)

    # Generate button
    st.divider()
    st.markdown("#### 📝 Final file name")
    final_name_input = st.text_input(
        "Type the name for the final masterfile (without extension)",
        value="final_masterfile",
        help="We'll add .xlsx or .xlsm automatically based on your template."
    )
    
    if st.button("🚀 Generate Final Masterfile", type="primary"):
        with st.spinner("🚀 Generating masterfile..."):
            # Rebuild mappings with user selections
            on_df = st.session_state.onboarding_data['df']
            series_by_alias = st.session_state.onboarding_data['series_by_alias']
            master_data = st.session_state.master_data
            used_cols = master_data['used_cols']
            display_headers = master_data['display_headers']
            secondary_headers = master_data['secondary_headers']
            master_bytes = master_data['bytes']
            ext = master_data['ext']
            
            BULLET_DISP_N = norm("Key Product Features")
            master_to_source = {}
            
            for c, (disp, sec) in enumerate(zip(display_headers, secondary_headers), start=1):
                disp_norm = norm(disp)
                sec_norm = norm(sec)
                if disp_norm == BULLET_DISP_N and sec_norm:
                    effective_header = sec
                else:
                    effective_header = disp
                eff_norm = norm(effective_header)
                if not eff_norm:
                    continue
                
                # Check user mappings first
                if c in st.session_state.user_mappings:
                    user_col = st.session_state.user_mappings[c]
                    if user_col in on_df.columns:
                        master_to_source[c] = on_df[user_col]
                        continue
                
                # Then check predefined mappings
                aliases = mapping_aliases.get(eff_norm, [effective_header])
                resolved = None
                for a in aliases:
                    s = series_by_alias.get(norm(a))
                    if s is not None:
                        resolved = s
                        break
                
                if resolved is not None:
                    master_to_source[c] = resolved
                elif disp_norm == norm("Listing Action (List or Unlist)"):
                    master_to_source[c] = SENTINEL_LIST
            
            n_rows = len(on_df)
            
            # Build data block
            block = [[""] * used_cols for _ in range(n_rows)]
            for col, src in master_to_source.items():
                if src is SENTINEL_LIST:
                    for i in range(n_rows):
                        block[i][col-1] = "List"
                else:
                    vals = src.astype(str).tolist()
                    m = min(len(vals), n_rows)
                    for i in range(m):
                        v = sanitize_xml_text(vals[i].strip())
                        if v and v.lower() not in ("nan", "none", ""):
                            block[i][col-1] = v
            
            # Fast XML write
            out_bytes = fast_patch_template(
                master_bytes=master_bytes,
                sheet_name=MASTER_TEMPLATE_SHEET,
                header_row=MASTER_DISPLAY_ROW,
                start_row=MASTER_DATA_START_ROW,
                used_cols=used_cols,
                block_2d=block
            )
            
            # Prepare download
            mime_map = {
                ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
            }
            out_mime = mime_map.get(ext, mime_map[".xlsx"])
            final_base = safe_filename(final_name_input, fallback="final_masterfile")
            final_filename = f"{final_base}{ext}"
            
            st.success("✅ Masterfile generated successfully!")
            st.download_button(
                "⬇️ Download Final Masterfile",
                data=out_bytes,
                file_name=final_filename,
                mime=out_mime,
                key="dl_final"
            )

with st.expander("📘 How to use (step-by-step)", expanded=False):
    st.markdown(dedent(f"""
    **This tool**
    - Comes with {len(PREDEFINED_MAPPING)} pre-configured attribute mappings for Amazon
    - Ultra-fast XML writer (generates files in seconds)
    - Preserves all sheets, styles, formulas, and macros (.xlsm support)

    **Steps**
    1) Upload the **Masterfile Template** (.xlsx/.xlsm) and **Onboarding** (.xlsx)
    2) Click **Analyze Files** to see which attributes are auto-mapped
    3) Map any remaining unmapped attributes using the dropdown selectors
    4) Choose your desired final file name
    5) Click **Generate Final Masterfile**
    6) Download your completed masterfile!
    """))
